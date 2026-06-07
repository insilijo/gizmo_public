"""Extended metadata discrimination — test every axis against ALL available
metadata fields per cohort (ordinal grade, continuous severity, subtype, etc.)
beyond just active/control.

Mines raw clinical files for cohorts where richer metadata is available.
"""
from __future__ import annotations

import gzip
import json
import re
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from scipy.stats import spearmanr
from sklearn.metrics import roc_auc_score

REPO = Path("/home/jgardner/GIZMO")
RESULTS = REPO / "benchmarks/results"
FIG_DIR = RESULTS / "figures"
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(REPO / "benchmarks"))


def load_kmplot_metadata():
    """KMPLOT_BRCA: Grade (ordinal 1-3), ER status, tumor_size."""
    try:
        candidates = [
            Path.home() / "gitlab-old/d2f483672c0239f6d7dd3c9ecee6deacbcd59185855625902a8b1c1a3bd67440/KMPLOT_BRCA_EXPRESSION/DATA/KMPLOT_BRCA_SURVIVAL.txt",
            Path.home() / ".cache" / "kmplot_brca" / "KMPLOT_BRCA_SURVIVAL.txt",
        ]
        path = next((p for p in candidates if p.exists()), None)
        if path is None:
            return None
        surv = pd.read_csv(path, sep="\t")
        surv["AffyID"] = surv["AffyID"].astype(str)
        cols = {"patient_id": surv["AffyID"].values}
        for c in ("Grade", "ER", "Size", "Node", "Age", "LymphNodeStatus"):
            if c in surv.columns:
                cols[c.lower()] = pd.to_numeric(surv[c], errors="coerce").values
        return pd.DataFrame(cols)
    except Exception as e:
        print(f"  KMPLOT metadata load failed: {e}", flush=True)
        return None


def _parse_stage_ordinal(v):
    """Parse Roman numeral or T-stage into ordinal 1-4."""
    s = str(v).upper().strip()
    if s in ("NA", "NAN", "", "UNKNOWN"): return np.nan
    if "IV" in s or "T4" in s or s.endswith("4"): return 4
    if "III" in s or "T3" in s or s.endswith("3"): return 3
    if ("II" in s and "III" not in s) or "T2" in s or s.endswith("2"): return 2
    if "I" in s and "II" not in s or "T1" in s or s.endswith("1"): return 1
    return np.nan


def _parse_grade_ordinal(v):
    s = str(v).upper().strip()
    if "G4" in s: return 4
    if "G3" in s: return 3
    if "G2" in s: return 2
    if "G1" in s: return 1
    return np.nan


def _add_field(meta, key, vals):
    """Coerce vals to numeric/binary and add to meta dict if non-constant."""
    arr = pd.Series(vals)
    # Try numeric
    num = pd.to_numeric(arr.replace(["NA", "nan", "Unknown", "unknown"], np.nan), errors="coerce")
    if num.notna().sum() >= 5 and len(num.dropna().unique()) >= 2:
        meta[key] = num.values
        return
    # Try binary string
    str_vals = arr.astype(str).str.strip().str.lower()
    uniq = set(str_vals[~str_vals.isin({"na", "nan", "unknown", ""})].unique())
    if uniq.issubset({"male", "female"}):
        meta[key] = [1 if v == "male" else 0 if v == "female" else np.nan for v in str_vals]
        return
    if uniq.issubset({"yes", "no"}):
        meta[key] = [1 if v == "yes" else 0 if v == "no" else np.nan for v in str_vals]
        return
    # Categorical → integer encoding
    cats = pd.Categorical(arr.astype(str).replace(["NA", "nan", "Unknown", "unknown", ""], np.nan))
    codes = cats.codes.astype(float)
    codes = np.where(codes < 0, np.nan, codes)
    if len(set(codes[~np.isnan(codes)])) >= 2:
        meta[key] = codes


def load_cptac_metadata(cohort):
    """CPTAC TSI: handles BOTH standard (rows=patients, cols=attrs, type-row at row 2)
    AND transposed (rows=attrs, cols=patients) formats."""
    name_map = {"CPTAC_CCRCC": "HS_CPTAC_CCRCC_CLI.tsi",
                "CPTAC_COAD": "HS_CPTAC_COAD_CLI.tsi",
                "CPTAC_OV": "HS_CPTAC_OV_CLI.tsi"}
    try:
        path = REPO / "data/cohorts" / cohort / name_map[cohort]
        if not path.exists():
            return None
        # Peek the first column header
        with open(path) as f:
            first_line = f.readline().strip()
        first_field = first_line.split("\t", 1)[0].lower()
        if first_field == "attrib_name":
            # TRANSPOSED format: rows = attributes, cols = patient IDs
            df = pd.read_csv(path, sep="\t")
            # df.columns: ['attrib_name', 'patient1', 'patient2', ...]
            attr_col = df.columns[0]
            df = df.set_index(attr_col)
            # Now df is (attributes × patients); transpose
            dft = df.T  # patients × attributes
            dft.index.name = "patient_id"
            dft = dft.reset_index()
            meta = {"patient_id": dft["patient_id"].astype(str).values}
            for c in dft.columns[1:]:
                cl_lc = c.lower()
                if cl_lc == "gender":
                    _add_field(meta, "sex", dft[c])
                elif cl_lc in ("stage", "pathology_t_stage", "pathology_n_stage"):
                    parsed = [_parse_stage_ordinal(v) for v in dft[c]]
                    if any(not np.isnan(v) for v in parsed):
                        meta[cl_lc] = parsed
                elif cl_lc == "age":
                    _add_field(meta, "age", dft[c])
                else:
                    _add_field(meta, cl_lc, dft[c])
            return pd.DataFrame(meta)
        else:
            # STANDARD format: rows = patients, first data row is "type" annotation, skip it
            df = pd.read_csv(path, sep="\t", skiprows=[1])
            df = df.rename(columns={df.columns[0]: "patient_id"})
            meta = {"patient_id": df["patient_id"].astype(str).values}
            for c in df.columns[1:]:
                cl = c.lower()
                # Handle CCRCC-style explicit columns
                if cl == "gender":
                    _add_field(meta, "sex", df[c])
                elif "age" in cl:
                    _add_field(meta, "age", df[c])
                elif "tumor_size" in cl:
                    _add_field(meta, "tumor_size_cm", df[c])
                elif cl == "bmi":
                    _add_field(meta, "bmi", df[c])
                elif "histologic_grade" in cl or cl == "grade":
                    parsed = [_parse_grade_ordinal(v) for v in df[c]]
                    if any(not np.isnan(v) for v in parsed):
                        meta["grade"] = parsed
                elif "stage" in cl:
                    parsed = [_parse_stage_ordinal(v) for v in df[c]]
                    if any(not np.isnan(v) for v in parsed):
                        meta[cl.replace(" ", "_").replace("(", "").replace(")", "")] = parsed
                else:
                    # Generic add (handles binary/categorical/numeric)
                    _add_field(meta, cl.replace(" ", "_").replace("(", "").replace(")", ""), df[c])
            return pd.DataFrame(meta)
    except Exception as e:
        print(f"  CPTAC {cohort} metadata load failed: {e}", flush=True)
        return None


def load_tcga_idh_glioma_metadata():
    """TCGA-IDH-glioma: IDH-mut status, grade, age, sex, codel, MGMT, transcriptome subtype."""
    try:
        path = Path.home() / ".cache" / "tcga_idh" / "lgggbm_tcga_pub_clinical.tsv"
        if not path.exists():
            return None
        df = pd.read_csv(path, sep="\t")
        # patient_id column — KEEP CASE so it matches F file uppercase TCGA IDs
        if "patient_id" not in df.columns:
            return None
        df["patient_id"] = df["patient_id"].astype(str)
        meta = {"patient_id": df["patient_id"].values}
        # Numeric
        for col in ("AGE", "KARNOFSKY_PERFORMANCE_SCORE", "OS_MONTHS"):
            if col in df.columns:
                meta[col.lower()] = pd.to_numeric(df[col], errors="coerce").values
        # Binary
        if "SEX" in df.columns:
            meta["sex"] = [1 if str(v).strip().lower() == "male" else
                            0 if str(v).strip().lower() == "female" else np.nan
                            for v in df["SEX"]]
        if "OS_STATUS" in df.columns:
            meta["os_event"] = [1 if "DECEASED" in str(v).upper() else
                                  0 if "LIVING" in str(v).upper() else np.nan
                                  for v in df["OS_STATUS"]]
        if "IDH_STATUS" in df.columns:
            meta["idh_mut"] = [1 if "MUT" in str(v).upper() else
                                0 if "WT" in str(v).upper() else np.nan
                                for v in df["IDH_STATUS"]]
        if "MGMT_PROMOTER_STATUS" in df.columns:
            meta["mgmt_methylated"] = [1 if "METH" in str(v).upper() and "UN" not in str(v).upper() else
                                       0 if "UNMETH" in str(v).upper() else np.nan
                                       for v in df["MGMT_PROMOTER_STATUS"]]
        # Ordinal grade (II/III/IV)
        if "GRADE" in df.columns:
            grade_map = {"G2": 2, "G3": 3, "G4": 4}
            meta["grade"] = [grade_map.get(str(v).strip().upper(), np.nan) for v in df["GRADE"]]
        # Categorical subtypes (encode as integer)
        for col in ("IDH_CODEL_SUBTYPE", "TRANSCRIPTOME_SUBTYPE", "ORIGINAL_SUBTYPE"):
            if col in df.columns:
                cats = pd.Categorical(df[col].astype(str).replace("nan", np.nan))
                meta[col.lower()] = cats.codes.astype(float)
                meta[col.lower()] = np.where(meta[col.lower()] < 0, np.nan, meta[col.lower()])
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  TCGA_IDH_glioma metadata load failed: {e}", flush=True)
        return None


def load_idh_glioma_trautwein_metadata():
    """Trautwein IDH-glioma (MTBLS3873 + GSE190504): IDH-mut status, histology, treatment."""
    try:
        path = REPO / "benchmarks/glioma_multiomics/sample_metadata.tsv"
        if not path.exists():
            return None
        df = pd.read_csv(path, sep="\t")
        # F file uses rna_id (e.g., '20087P01A01_01'); metadata table has rna_id column
        if "rna_id" in df.columns:
            df["patient_id"] = df["rna_id"].astype(str)
        elif "description" in df.columns:
            df["patient_id"] = df["description"].astype(str).str.replace(r"_(NMR|RNA)$", "", regex=True)
        else:
            return None
        meta = {"patient_id": df["patient_id"].values}
        if "genotype" in df.columns:
            meta["idh_mut"] = [1 if "mut" in str(v).lower() else
                                0 if "wt" in str(v).lower() else np.nan
                                for v in df["genotype"]]
        if "histology" in df.columns:
            cats = pd.Categorical(df["histology"].astype(str).replace("nan", np.nan))
            meta["histology"] = cats.codes.astype(float)
            meta["histology"] = np.where(meta["histology"] < 0, np.nan, meta["histology"])
        if "treatment" in df.columns:
            meta["treatment_received"] = [0 if "untreated" in str(v).lower() else
                                            1 if "treated" in str(v).lower() else np.nan
                                            for v in df["treatment"]]
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  IDH-glioma Trautwein metadata load failed: {e}", flush=True)
        return None


def load_gao_ra_metadata():
    """Gao 2020 RA-vs-OA: GSE100786 microarray sample chars + MTBLS564 NMR sample sheet.
    Returns combined metadata keyed by sample ID."""
    try:
        # GSE100786 series_matrix has !Sample_characteristics_ch1 with disease/age/sex/etc
        ra_dir = REPO / "benchmarks" / "ra_multiomics"
        gse_path = ra_dir / "GSE100786_series_matrix.txt.gz"
        if not gse_path.exists():
            gse_path = ra_dir / "GSE100786_series_matrix.txt"
        gse_meta = load_gse_series_metadata(gse_path) if gse_path.exists() else None
        # MTBLS564 sample sheet
        s_path = ra_dir / "s_MTBLS564.txt"
        mtbls_meta = None
        if s_path.exists():
            try:
                s_df = pd.read_csv(s_path, sep="\t")
                pid_col = next((c for c in s_df.columns if "sample name" in c.lower() or "source" in c.lower()), None)
                if pid_col:
                    s_df["patient_id"] = s_df[pid_col].astype(str)
                    keep = {"patient_id": s_df["patient_id"].values}
                    for c in s_df.columns:
                        cl = c.lower()
                        if "disease" in cl or "diagnosis" in cl:
                            cats = pd.Categorical(s_df[c].astype(str).replace("nan", np.nan))
                            keep["disease_mtbls"] = cats.codes.astype(float)
                            keep["disease_mtbls"] = np.where(keep["disease_mtbls"] < 0, np.nan, keep["disease_mtbls"])
                        elif "age" in cl:
                            keep["age_mtbls"] = pd.to_numeric(s_df[c], errors="coerce").values
                        elif "sex" in cl or "gender" in cl:
                            keep["sex_mtbls"] = [1 if str(v).lower() == "male" else
                                                   0 if str(v).lower() == "female" else np.nan
                                                   for v in s_df[c]]
                    mtbls_meta = pd.DataFrame(keep)
            except Exception:
                pass
        # Return either GSE meta or MTBLS meta (preferring GSE)
        return gse_meta if gse_meta is not None and len(gse_meta) > 0 else mtbls_meta
    except Exception as e:
        print(f"  Gao_RA metadata load failed: {e}", flush=True)
        return None


def load_crohn_metadata():
    """Koopman 2025 pediatric Crohn: disease state, thiopurine treatment."""
    try:
        import openpyxl
        path = REPO / "benchmarks" / "nci60" / "crohn_supp4.xlsx"
        if not path.exists():
            return None
        wb = openpyxl.load_workbook(path, read_only=True)
        ws1 = wb["1"]
        rows = list(ws1.iter_rows(values_only=True))
        # header: ('PIN subject', 'Disease state based on calprotectin', 'Thiopurine')
        pids, thio = [], []
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            pids.append(str(row[0]).strip())
            thio.append(1 if str(row[2]).strip().lower() == "yes" else
                          0 if str(row[2]).strip().lower() == "no" else np.nan)
        return pd.DataFrame({"patient_id": pids, "thiopurine": thio})
    except Exception as e:
        print(f"  Crohn metadata load failed: {e}", flush=True)
        return None


def load_filbin_metadata():
    """Filbin MGH-COVID: WHO acuity (continuous), comorbidities, labs, age/BMI cat."""
    try:
        path = Path.home() / ".cache" / "filbin_mgh_covid" / "Clinical_Metadata.xlsx"
        if not path.exists():
            return None
        df = pd.read_excel(path)
        df["patient_id"] = df["Public ID"].astype(int).astype(str)
        meta = {"patient_id": df["patient_id"].values}
        # Numeric / ordinal fields
        keep_numeric = ["COVID", "Age cat", "BMI cat", "Acuity 0", "Acuity 3", "Acuity 7",
                         "Acuity 28", "Acuity max",
                         "abs_neut_0_cat", "abs_lymph_0_cat", "abs_mono_0_cat",
                         "creat_0_cat", "crp_0_cat", "ddimer_0_cat", "ldh_0_cat", "Trop_72h"]
        for c in keep_numeric:
            if c in df.columns:
                key = c.lower().replace(" ", "_")
                meta[key] = pd.to_numeric(df[c], errors="coerce").values
        # Binary symptom flags
        for c in ["HEART", "LUNG", "KIDNEY", "DIABETES", "HTN", "IMMUNO",
                   "Resp_Symp", "Fever_Sympt", "GI_Symp"]:
            if c in df.columns:
                meta[c.lower()] = pd.to_numeric(df[c], errors="coerce").values
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  Filbin metadata load failed: {e}", flush=True)
        return None


def load_erawijantari_metadata():
    """Erawijantari gastric cancer: study group, surgery type, age, gender, BMI, comorbidities."""
    try:
        path = Path.home() / ".cache" / "gemma" / "demo" / "curated" / "ERAWIJANTARI_GASTRIC_CANCER_2020" / "metadata.tsv"
        if not path.exists():
            return None
        df = pd.read_csv(path, sep="\t")
        df["patient_id"] = df["Sample"].astype(str)
        meta = {"patient_id": df["patient_id"].values}
        if "Age" in df.columns:
            meta["age"] = pd.to_numeric(df["Age"], errors="coerce").values
        if "Gender" in df.columns:
            meta["sex"] = [1 if str(v).lower() == "male" else
                            0 if str(v).lower() == "female" else np.nan
                            for v in df["Gender"]]
        if "BMI" in df.columns:
            meta["bmi"] = pd.to_numeric(df["BMI"], errors="coerce").values
        if "Study.Group" in df.columns:
            cats = pd.Categorical(df["Study.Group"].astype(str).replace("nan", np.nan))
            meta["study_group"] = cats.codes.astype(float)
            meta["study_group"] = np.where(meta["study_group"] < 0, np.nan, meta["study_group"])
        if "Surgery_Type" in df.columns:
            cats = pd.Categorical(df["Surgery_Type"].astype(str).replace("nan", np.nan))
            meta["surgery_type"] = cats.codes.astype(float)
            meta["surgery_type"] = np.where(meta["surgery_type"] < 0, np.nan, meta["surgery_type"])
        # Comorbidity flags — Yes/No → 1/0
        for c in ["Hypertension", "Diabetes", "Dyslipidemia", "DiabetesMed",
                   "Anticoagulant", "Analgesic"]:
            if c in df.columns:
                meta[c.lower()] = [1 if str(v).strip().lower() == "yes" else
                                     0 if str(v).strip().lower() == "no" else np.nan
                                     for v in df[c]]
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  Erawijantari metadata load failed: {e}", flush=True)
        return None


def load_hmp2_metadata():
    """HMP2 IBD-CD: diagnosis, sex, age, HBI score, location."""
    try:
        path = Path.home() / "GeMMA" / "data" / "hmp2" / "hmp2_sample_metadata.csv"
        if not path.exists():
            return None
        df = pd.read_csv(path)
        # ID column: try sample_id, External ID, Participant ID
        pid_col = None
        for cand in ("sample_id", "External ID", "Participant ID"):
            if cand in df.columns:
                pid_col = cand
                break
        if pid_col is None:
            return None
        df["patient_id"] = df[pid_col].astype(str)
        meta = {"patient_id": df["patient_id"].values}
        if "consent_age" in df.columns:
            meta["age"] = pd.to_numeric(df["consent_age"], errors="coerce").values
        elif "Age at diagnosis" in df.columns:
            meta["age"] = pd.to_numeric(df["Age at diagnosis"], errors="coerce").values
        if "sex" in df.columns:
            meta["sex"] = [1 if str(v).lower() == "male" else
                            0 if str(v).lower() == "female" else np.nan
                            for v in df["sex"]]
        for col in ("hbi", "HBI", "HBI_Score"):
            if col in df.columns:
                meta["hbi"] = pd.to_numeric(df[col], errors="coerce").values
                break
        for col in ("Location at diagnosis (Montreal)", "site_name"):
            if col in df.columns:
                cats = pd.Categorical(df[col].astype(str).replace("nan", np.nan))
                meta[col.lower().replace(" ", "_")] = cats.codes.astype(float)
                meta[col.lower().replace(" ", "_")] = np.where(
                    meta[col.lower().replace(" ", "_")] < 0, np.nan,
                    meta[col.lower().replace(" ", "_")]
                )
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  HMP2 metadata load failed: {e}", flush=True)
        return None


def load_tcga_luad_metadata():
    """TCGA_LUAD: stage (ordinal 1-4)."""
    try:
        path = Path.home() / ".cache" / "tcga_luad" / "gdac.broadinstitute.org_LUAD.Clinical_Pick_Tier1.Level_4.2016012800.0.0" / "LUAD.clin.merged.picked.txt"
        if not path.exists():
            return None
        cdf = pd.read_csv(path, sep="\t", header=None, low_memory=False)
        attrs = cdf.iloc[:, 0].astype(str).tolist()
        pids = [str(p).strip().lower() for p in cdf.iloc[0, 1:].tolist()]
        meta = {"patient_id": pids}
        def _parse_stage(s):
            if not isinstance(s, str): return np.nan
            s = s.lower().strip()
            if "iv" in s: return 4
            if "iii" in s: return 3
            if "ii" in s and "iii" not in s: return 2
            if "i" in s and "ii" not in s and "iii" not in s: return 1
            return np.nan
        if "pathologic_stage" in attrs:
            idx = attrs.index("pathologic_stage")
            stages = cdf.iloc[idx, 1:].tolist()
            meta["stage"] = [_parse_stage(s) for s in stages]
        for field in ("age_at_initial_pathologic_diagnosis", "gender", "histological_type"):
            if field in attrs:
                idx = attrs.index(field)
                vals = cdf.iloc[idx, 1:].tolist()
                if field == "gender":
                    meta["sex"] = [1 if str(v).lower().strip() == "male" else 0 if str(v).lower().strip() == "female" else np.nan
                                    for v in vals]
                else:
                    meta[field] = [pd.to_numeric(v, errors="coerce") for v in vals]
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  tcga_luad metadata load failed: {e}", flush=True)
        return None


def load_su_covid_metadata():
    """Su_COVID: WHO Ordinal Scale, age, sex, BMI, comorbidities — from Table S1.1."""
    try:
        import openpyxl
        path = REPO / "benchmarks" / ("Table S1. Human subject details, plasma proteomic "
                                       "and metabolomic datasets and analysis, and CITE-seq "
                                       "antibodies. Related to Figures 1 and S1.xlsx")
        if not path.exists():
            return None
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["S1.1 Patient Clinical Data"]
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[0])
        # Index columns of interest
        col_idx = {h: i for i, h in enumerate(header) if h}
        meta_rows = []
        for r in rows[1:]:
            if not r or not r[col_idx.get("Sample ID", 1)]:
                continue
            meta_rows.append({
                "patient_id": str(r[col_idx["Sample ID"]]),
                "who_ordinal_scale": pd.to_numeric(r[col_idx.get("Who Ordinal Scale", 4)], errors="coerce"),
                "sex": (1 if str(r[col_idx.get("Sex", 5)]).lower() == "male" else
                         0 if str(r[col_idx.get("Sex", 5)]).lower() == "female" else np.nan),
                "age": pd.to_numeric(r[col_idx.get("Age", 6)], errors="coerce"),
                "bmi": pd.to_numeric(r[col_idx.get("BMI", 9)], errors="coerce"),
            })
            # Add comorbidities if present (Yes/No → 1/0)
            for clin in ("Asthma", "Cancer", "Chronic Hypertension", "Chronic Kidney Disease"):
                if clin in col_idx:
                    v = r[col_idx[clin]]
                    meta_rows[-1][clin.lower().replace(" ", "_")] = (
                        1 if str(v).strip().lower() == "yes" else
                        0 if str(v).strip().lower() == "no" else np.nan
                    )
        # Also add healthy donors with sex/age only (separate sheet)
        try:
            ws_h = wb["S1.2 Healthy Donor Info"]
            rows_h = list(ws_h.iter_rows(values_only=True))
            h_hdr = rows_h[0]
            for r in rows_h[1:]:
                if not r or not r[0]:
                    continue
                meta_rows.append({
                    "patient_id": str(r[0]),
                    "sex": (1 if str(r[1]).lower() == "male" else
                             0 if str(r[1]).lower() == "female" else np.nan),
                    "age": pd.to_numeric(r[2], errors="coerce"),
                    "who_ordinal_scale": 0,  # healthy by definition
                })
        except Exception:
            pass
        return pd.DataFrame(meta_rows)
    except Exception as e:
        print(f"  Su_COVID metadata load failed: {e}", flush=True)
        return None


def load_su_covid_metadata_legacy():
    """Legacy CSV-based loader (kept for fallback)."""
    try:
        path = Path.home() / ".cache" / "su_covid" / "clinical.csv"
        if not path.exists():
            return None
        df = pd.read_csv(path)
        meta = {}
        if "sample_id" in df.columns:
            df["patient_id"] = df["sample_id"].astype(str).str.split("-").str[0]
        elif "subject" in df.columns:
            df["patient_id"] = df["subject"].astype(str)
        else:
            return None
        meta["patient_id"] = df["patient_id"].values
        for cand in ("acuity", "Acuity", "WHO_max", "who_max"):
            if cand in df.columns:
                meta["acuity"] = pd.to_numeric(df[cand], errors="coerce").values
                break
        for cand in ("age", "Age"):
            if cand in df.columns:
                meta["age"] = pd.to_numeric(df[cand], errors="coerce").values
                break
        for cand in ("sex", "Sex", "gender"):
            if cand in df.columns:
                meta["sex"] = [1 if str(v).lower().strip() in ("male", "m") else
                                0 if str(v).lower().strip() in ("female", "f") else np.nan
                                for v in df[cand].values]
                break
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  Su_COVID metadata load failed: {e}", flush=True)
        return None


def load_gse_series_metadata(geo_file: Path):
    """Generic GEO series_matrix metadata extractor."""
    try:
        opener = gzip.open if str(geo_file).endswith(".gz") else open
        sample_titles = []
        characteristics = {}
        with opener(geo_file, "rt", errors="ignore") as f:
            for line in f:
                if line.startswith("!Sample_geo_accession"):
                    sample_titles = [s.strip().strip('"') for s in line.strip().split("\t")[1:]]
                elif line.startswith("!Sample_characteristics_ch1"):
                    vals = [s.strip().strip('"') for s in line.strip().split("\t")[1:]]
                    if vals and ":" in vals[0]:
                        key = vals[0].split(":")[0].strip()
                        key = re.sub(r"[^a-zA-Z0-9_]+", "_", key).lower()
                        characteristics[key] = [v.split(":", 1)[1].strip() if ":" in v else v
                                                  for v in vals]
                elif line.startswith("!series_matrix_table_begin"):
                    break
        if not sample_titles:
            return None
        meta = {"patient_id": sample_titles}
        for key, vals in characteristics.items():
            num_series = pd.Series(pd.to_numeric(vals, errors="coerce"))
            if num_series.notna().sum() >= len(vals) * 0.5:
                meta[key] = num_series.values
            else:
                cnts = pd.Series(vals).value_counts()
                if len(cnts) == 2:
                    most_common = cnts.idxmax()
                    meta[key] = [0 if v == most_common else 1 if v and not pd.isna(v) else np.nan
                                  for v in vals]
                elif 2 < len(cnts) <= 10:
                    # Multi-class ordinal-like → assign ranks by frequency (or alphabetical)
                    cat_map = {c: i for i, c in enumerate(sorted(cnts.index))}
                    meta[key] = [cat_map.get(v, np.nan) for v in vals]
        return pd.DataFrame(meta)
    except Exception as e:
        print(f"  GEO metadata load failed for {geo_file}: {e}", flush=True)
        return None


def extract_metadata_for_cohort(cohort, patient_ids):
    """Returns DataFrame keyed by patient_id with available metadata columns."""
    if cohort == "KMPLOT_BRCA":
        return load_kmplot_metadata()
    if cohort == "TCGA_LUAD":
        return load_tcga_luad_metadata()
    if cohort == "TCGA_IDH_glioma":
        return load_tcga_idh_glioma_metadata()
    if cohort == "IDH_glioma":
        return load_idh_glioma_trautwein_metadata()
    if cohort == "HMP2_IBD_CD":
        return load_hmp2_metadata()
    if cohort == "Filbin_COVID":
        return load_filbin_metadata()
    if cohort == "Erawijantari":
        return load_erawijantari_metadata()
    if cohort == "Crohn":
        return load_crohn_metadata()
    if cohort == "Gao_RA":
        return load_gao_ra_metadata()
    if cohort.startswith("CPTAC_"):
        return load_cptac_metadata(cohort)
    if cohort == "Su_COVID":
        return load_su_covid_metadata()
    if cohort == "GSE65682_sepsis":
        return load_gse_series_metadata(
            REPO / "data/cohorts/GSE65682_sepsis/GSE65682_series_matrix.txt.gz")
    if cohort == "GSE65391_SLE":
        return load_gse_series_metadata(
            REPO / "data/cohorts/GSE65391_SLE/GSE65391_series_matrix.txt.gz")
    if cohort == "GSE89408_RA":
        cache_path = Path.home() / ".cache" / "gse89408" / "GSE89408_series_matrix.txt.gz"
        if cache_path.exists():
            return load_gse_series_metadata(cache_path)
        return None
    return None


def discriminate(axis_vals, meta_vals):
    try:
        arr = np.asarray(meta_vals, dtype=np.float64)
    except (TypeError, ValueError):
        return ("uncoercible_meta", float("nan"))
    axis_vals = np.asarray(axis_vals, dtype=np.float64)
    mask = ~np.isnan(arr) & ~np.isnan(axis_vals)
    if mask.sum() < 6:
        return ("insufficient_n", float("nan"))
    a = axis_vals[mask]; m = arr[mask]
    uniq = np.unique(m)
    if len(uniq) < 2:
        return ("constant_meta", float("nan"))
    if len(uniq) == 2:
        try:
            auc = float(roc_auc_score(m, np.abs(a)))
            return ("auc", auc)
        except Exception:
            return ("auc_failed", float("nan"))
    elif len(uniq) <= 10:
        try:
            rho, _ = spearmanr(a, m)
            return ("spearman", float(abs(rho)))
        except Exception:
            return ("spearman_failed", float("nan"))
    else:
        try:
            rho, _ = spearmanr(a, m)
            return ("spearman", float(abs(rho)))
        except Exception:
            return ("spearman_failed", float("nan"))


def main():
    df_axis = pd.read_csv(RESULTS / "n1_held_out_inference.tsv", sep="\t")
    pc_cols = ["beta_z", "alpha_pc1_z", "alpha_pc2_z", "alpha_pc3_z",
               "alpha_pc4_z", "alpha_pc5_z", "alpha_norm_z"]
    pc_labels = ["β", "α-PC1", "α-PC2", "α-PC3", "α-PC4", "α-PC5", "‖α‖₂"]

    # Add active/control as baseline metadata
    df_axis["active_vs_control"] = df_axis["label_bin"].replace(-1, np.nan)

    rows = []
    for cohort in sorted(df_axis["cohort"].unique()):
        sub = df_axis[df_axis["cohort"] == cohort].reset_index(drop=True)
        if len(sub) < 10:
            continue
        print(f"\n=== {cohort} ===", flush=True)
        # Start with active_vs_control baseline
        meta_df = pd.DataFrame({"patient_id": sub["patient_id"].astype(str),
                                  "active_vs_control": sub["active_vs_control"].values})
        # Add cohort-specific metadata; for CPTAC, strip _T/_N suffix on F-side IDs
        extra = extract_metadata_for_cohort(cohort, sub["patient_id"].tolist())
        if extra is not None and len(extra) > 0:
            extra["patient_id"] = extra["patient_id"].astype(str)
            if cohort.startswith("CPTAC_"):
                meta_df["patient_id_base"] = meta_df["patient_id"].str.replace(r"_[TN]$", "", regex=True)
                meta_df = meta_df.merge(
                    extra.rename(columns={"patient_id": "patient_id_base"}),
                    on="patient_id_base", how="left", suffixes=("", "_x")
                ).drop(columns=["patient_id_base"], errors="ignore")
            else:
                meta_df = meta_df.merge(extra, on="patient_id", how="left", suffixes=("", "_x"))
            print(f"  loaded extra metadata: {[c for c in extra.columns if c != 'patient_id']}",
                  flush=True)

        merged = sub.copy()
        merged["patient_id"] = merged["patient_id"].astype(str)
        merged = merged.merge(meta_df, on="patient_id", how="left", suffixes=("", "_meta"))
        meta_cols = [c for c in meta_df.columns
                     if c != "patient_id" and merged[c].notna().sum() >= 6]
        if not meta_cols:
            print(f"  no usable metadata", flush=True)
            continue
        for axis_col, axis_lbl in zip(pc_cols, pc_labels):
            for meta_col in meta_cols:
                metric_name, val = discriminate(
                    merged[axis_col].values, merged[meta_col].values
                )
                rows.append({
                    "cohort": cohort, "axis": axis_lbl, "metadata": meta_col,
                    "metric": metric_name, "value": val,
                    "n": int(merged[meta_col].notna().sum()),
                })

    out_df = pd.DataFrame(rows)
    out_tsv = RESULTS / "axis_metadata_discrimination_extended.tsv"
    out_df.to_csv(out_tsv, sep="\t", index=False)
    print(f"\nWrote {out_tsv} ({len(out_df)} tests)", flush=True)

    # Print top discrimination per cohort × metadata
    print(f"\n{'='*80}\nTop axis per cohort × metadata\n{'='*80}", flush=True)
    for cohort in out_df["cohort"].unique():
        c_df = out_df[out_df["cohort"] == cohort]
        meta_cols = c_df["metadata"].unique()
        for meta in meta_cols:
            m_df = c_df[c_df["metadata"] == meta]
            if m_df["value"].notna().sum() == 0: continue
            top = m_df.nlargest(2, "value")[["axis", "value", "metric", "n"]]
            top_str = ", ".join([f"{r['axis']}={r['value']:.2f}({r['metric']})" for _, r in top.iterrows()])
            n = int(m_df["n"].iloc[0])
            print(f"  {cohort:<22s} | {meta:<28s} | n={n:<4d} | top: {top_str}",
                  flush=True)

    # Per-cohort heatmap for each metadata field
    valid_cohorts = sorted(out_df["cohort"].unique())
    valid_metas = sorted(set(out_df["metadata"].dropna().unique()))
    if len(valid_metas) <= 1:
        return
    # Long-form table: rows = cohort×metadata, cols = axis
    long_df = out_df.pivot_table(
        index=["cohort", "metadata"], columns="axis", values="value", aggfunc="first"
    )
    fig, ax = plt.subplots(figsize=(9, max(5, 0.35 * len(long_df))))
    arr = long_df.values
    im = ax.imshow(arr, cmap="viridis", aspect="auto", vmin=0.3, vmax=1.0)
    ax.set_xticks(range(len(long_df.columns)))
    ax.set_xticklabels(long_df.columns, rotation=45, ha="right", fontsize=9)
    ax.set_yticks(range(len(long_df)))
    ax.set_yticklabels([f"{c} | {m}" for c, m in long_df.index], fontsize=7)
    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            val = arr[i, j]
            if not np.isnan(val):
                text_color = "white" if val > 0.7 or val < 0.4 else "black"
                ax.text(j, i, f"{val:.2f}", ha="center", va="center",
                        fontsize=6, color=text_color)
    plt.colorbar(im, ax=ax, label="|discrimination| (AUC binary / |ρ| ordinal/continuous)")
    ax.set_title("Per-axis discrimination against every available metadata field\n"
                 "(no axis selection — every axis tested against every cohort's metadata)",
                 fontsize=10)
    plt.tight_layout()
    out_png = FIG_DIR / "fig_axis_metadata_extended.png"
    plt.savefig(out_png, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"\nWrote {out_png}", flush=True)


if __name__ == "__main__":
    main()
