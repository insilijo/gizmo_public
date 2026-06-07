"""Master per-patient benchmark — runs submodular + WL/SP graph kernels
across cohorts × designs.

Designs:
  edge-informed:    trans/prot evidence only (gene → enzyme nodes → reaction edges)
  node-informed:    metab evidence only (substrate/product nodes)
  both_directional: concat both with per-omic L2 normalization

Cohorts:
  Crohn (n≈33, prot+metab)            — all 3 designs
  Su COVID (n≈270, prot+metab)         — all 3 designs
  Erawijantari (n≈96, metab only)      — node-informed only

Each (cohort, design, kernel) cell reports:
  CV AUC ± SD (5-fold stratified, class-balanced SVC)
  LOO AUC
  Hold-out AUC (70/30 stratified)

Output:
  benchmarks/results/per_patient_master.tsv
"""
from __future__ import annotations
import sys, math, statistics
from pathlib import Path
from collections import Counter

import numpy as np

REPO = Path(__file__).resolve().parent.parent
RESULTS = REPO / "benchmarks" / "results"
sys.path.insert(0, str(REPO / "benchmarks"))


# ---------------------------------------------------------------------------
# Cohort loaders → return (prot_data, metab_data, group_labels)
# ---------------------------------------------------------------------------

def load_crohn():
    from per_patient_classification import load_crohn_matrices
    prot_data, metab_raw, groups = load_crohn_matrices()
    # Log-transform metab; strip biofluid suffix ("_plasma" etc.) so feature
    # names match canonical chemical names in MetaboliteMapper (which knows
    # "Arginine", "ATP", "Glucose-6P" but not "Arginine_plasma").
    common = sorted(set(prot_data) & set(metab_raw) & set(groups))
    _BIOFLUID_SUFFIXES = ("_plasma", "_serum", "_urine", "_csf", "_feces",
                          "_stool", "_saliva", "_synovial")
    def _strip(name: str) -> str:
        n = name.strip()
        for suf in _BIOFLUID_SUFFIXES:
            if n.lower().endswith(suf):
                return n[: -len(suf)]
        return n
    metab_log = {}
    for s in common:
        d = {}
        for k, v in metab_raw[s].items():
            if v > 0:
                key = _strip(k)
                d[key] = math.log2(v + 1e-9)
        metab_log[s] = d
    y_label = {s: ("active" if groups[s] == "active" else "control") for s in common}
    return prot_data, metab_log, y_label, common


def load_su_covid():
    """Same loader logic as per_patient_su_covid.py main()."""
    import openpyxl
    from laplacian_su_covid import map_descriptions_to_genes

    DATA_CANDIDATES = [
        Path("/root/.cache/su_covid"),
        Path.home() / ".cache" / "su_covid",
        REPO / "benchmarks" / "su_covid",
        REPO / "benchmarks",
    ]
    SUPP = None
    for d in DATA_CANDIDATES:
        try:
            if not d.exists():
                continue
        except PermissionError:
            continue
        for p in d.glob("*.xlsx"):
            n = p.name.lower()
            if any(k in n for k in ("su", "covid", "human subject details", "plasma proteomic")):
                SUPP = p
                break
        if SUPP is not None:
            break
    if SUPP is None:
        raise FileNotFoundError("Su 2020 supplementary tables not found")

    wb = openpyxl.load_workbook(SUPP, read_only=True)

    def covid_label(grp: str):
        if grp == "COVID19": return "active"
        if grp == "Healthy Donor": return "control"
        return None

    ws3 = wb["S1.3 Plasma Proteomic Data"]
    rows3 = list(ws3.iter_rows(values_only=True))
    descriptions = list(rows3[0][2:])
    desc_to_sym = map_descriptions_to_genes(descriptions)

    prot_data, prot_groups = {}, {}
    for row in rows3[1:]:
        if row[1] is None: continue
        grp = str(row[0]).strip() if row[0] else ""
        sid = str(row[1])
        label = covid_label(grp)
        if label is None: continue
        prot_groups[sid] = label
        d = {}
        for desc, v in zip(descriptions, row[2:]):
            sym = desc_to_sym.get(desc)
            if not sym or v is None or v == "": continue
            try: d[sym] = float(v)
            except (TypeError, ValueError): continue
        prot_data[sid] = d

    ws4 = wb["S1.4 Plasma Metabolomic Data"]
    rows4 = list(ws4.iter_rows(values_only=True))
    metab_names = [str(c).strip() for c in rows4[0][2:] if c]

    metab_data, metab_groups = {}, {}
    for row in rows4[1:]:
        if row[1] is None: continue
        grp = str(row[0]).strip() if row[0] else ""
        sid = str(row[1])
        label = covid_label(grp)
        if label is None: continue
        metab_groups[sid] = label
        d = {}
        for name, v in zip(metab_names, row[2:]):
            if v is None or v == "": continue
            try: d[name] = float(v)
            except (TypeError, ValueError): continue
        metab_data[sid] = d

    common = sorted(set(prot_data) & set(metab_data) & set(prot_groups))
    metab_log = {s: {k: math.log2(v + 1e-9) for k, v in metab_data[s].items() if v > 0}
                  for s in common}
    y_label = {s: prot_groups[s] for s in common}
    return prot_data, metab_log, y_label, common


def load_gao_ra():
    """Load Gao 2020 RA-vs-OA: GSE100786 microarray (per-sample) + MTBLS564 NMR (per-sample).

    Note cohorts are not individually-paired: GSM IDs (microarray) and
    sample numbers 1-24 (NMR) come from different patient sets. We load
    them with disjoint keysets; only edge_informed and node_informed
    designs are evaluated (both_directional skipped — no shared IDs).

    Returns (trans_data, metab_data, y_label, common) where:
      trans_data keyed by GSM ID
      metab_data keyed by sample number (str)
      common is the union (per-design subset is filtered in run_cell)
    """
    import csv, gzip
    BENCH = REPO / "benchmarks" / "ra_multiomics"

    # --- GSE100786 microarray: probe × sample, then aggregate to gene level ---
    SERIES = BENCH / "GSE100786_series_matrix.txt.gz"
    probe_to_gene: dict[str, str] = {}
    with open(BENCH / "GPL570_table.txt") as f:
        in_table = False
        header = None
        for line in f:
            line = line.rstrip("\n")
            if line == "!platform_table_begin":
                in_table = True
                continue
            if not in_table:
                continue
            if line == "!platform_table_end":
                break
            if header is None:
                header = line.split("\t")
                continue
            parts = line.split("\t")
            if len(parts) < 11:
                continue
            probe, gene = parts[0], parts[10]
            if gene and gene != "---":
                # Take first symbol if multi-mapped (e.g. "DDR1 /// MIR4640")
                gene = gene.split(" /// ")[0].strip()
                probe_to_gene[probe] = gene

    sample_diag: dict[str, str] = {}
    sample_ids: list[str] = []
    trans_data: dict[str, dict[str, float]] = {}

    with gzip.open(SERIES, "rt") as f:
        diag_line = None
        in_table = False
        header_line = None
        gene_sums: dict[str, dict[str, float]] = {}
        gene_counts: dict[str, dict[str, int]] = {}
        for line in f:
            line = line.rstrip("\n")
            if line.startswith("!Sample_title"):
                titles = [s.strip('"') for s in line.split("\t")[1:]]
            elif line.startswith("!Series_sample_id"):
                sample_ids = line.split("\t")[1].strip('"').split()
            elif line.startswith("!Sample_characteristics_ch1") and "diagnosis:" in line:
                diags = [s.strip('"').replace("diagnosis: ", "") for s in line.split("\t")[1:]]
                for sid, d in zip(sample_ids, diags):
                    if "Rheumatoid" in d:
                        sample_diag[sid] = "active"
                    elif "Osteo" in d:
                        sample_diag[sid] = "control"
            elif line == "!series_matrix_table_begin":
                in_table = True
                continue
            elif line == "!series_matrix_table_end":
                break
            elif in_table:
                if header_line is None:
                    header_line = [s.strip('"') for s in line.split("\t")]
                    for sid in header_line[1:]:
                        gene_sums[sid] = {}
                        gene_counts[sid] = {}
                    continue
                parts = line.split("\t")
                probe = parts[0].strip('"')
                gene = probe_to_gene.get(probe)
                if not gene:
                    continue
                for sid, val in zip(header_line[1:], parts[1:]):
                    try:
                        v = float(val)
                    except ValueError:
                        continue
                    gene_sums[sid][gene] = gene_sums[sid].get(gene, 0.0) + v
                    gene_counts[sid][gene] = gene_counts[sid].get(gene, 0) + 1

        # Aggregate: mean of probe values per gene per sample
        for sid in sample_ids:
            gene_avg = {}
            for g, total in gene_sums.get(sid, {}).items():
                gene_avg[g] = total / max(1, gene_counts[sid].get(g, 1))
            trans_data[sid] = gene_avg

    print(f"  Gao trans samples: {Counter(sample_diag.values())} ({len(trans_data)})")

    # --- MTBLS564 NMR: metabolite × sample, with RA/OA labels from sample sheet ---
    metab_data: dict[str, dict[str, float]] = {}
    metab_diag: dict[str, str] = {}
    SAMPLE_META = BENCH / "s_MTBLS564.txt"
    with SAMPLE_META.open() as f:
        for row in csv.DictReader(f, delimiter="\t"):
            sid = str(row["Sample Name"]).strip()
            d = row["Factor Value[Diagnosis]"].strip()
            if d == "RA":
                metab_diag[sid] = "active"
            elif d == "OA":
                metab_diag[sid] = "control"

    MAF_TSV = BENCH / "m_MTBLS564_metabolite_profiling_NMR_spectroscopy_v2_maf.tsv"
    with MAF_TSV.open() as f:
        reader = csv.DictReader(f, delimiter="\t")
        sample_cols = [c for c in reader.fieldnames
                        if c.strip().isdigit() and c.strip() in metab_diag]
        for row in reader:
            name = row.get("metabolite_identification", "").strip()
            chebi = row.get("database_identifier", "").strip()
            if not name or name == "unknown":
                continue
            # Prefer ChEBI ID for mapping (cleaner than name-fuzzy);
            # fall back to name when ChEBI is empty.
            key = chebi if chebi else name
            for sid in sample_cols:
                v = row.get(sid, "").strip()
                if not v:
                    continue
                try:
                    val = float(v)
                except ValueError:
                    continue
                if val <= 0 or math.isnan(val):
                    continue
                metab_data.setdefault(sid, {})[key] = math.log2(val + 1e-9)

    print(f"  Gao metab samples: {Counter(metab_diag.values())} ({len(metab_data)})")

    # Gao_RA is NOT actually paired multi-omic: GSM IDs (microarray) and
    # sample numbers 1-24 (NMR) come from different patient sets per the
    # source data. The union preserves both groups; Stage 3 reconstruction
    # uses a per-patient anchor (only adds diag weight at nodes this
    # patient actually observed), so missing-modality patients get
    # reconstructions driven by their single available modality without
    # spurious zero-anchors at the absent modality's nodes.
    y_label = {**sample_diag, **metab_diag}
    common = sorted(set(trans_data.keys()) | set(metab_data.keys()))
    return trans_data, metab_data, y_label, common


def load_corevitas():
    """Load CorEvitas RA cohort (V0 baseline trans-only) using the metabolism-
    curated 1,220-gene panel from `gitlab-old/.../corevitas_gse129705_amdbnorm_combined.exp_data.txt`.

    Why this panel: the standard `corevitas_rnaseq_vst_symbols.csv` (942 genes,
    immune-curated) maps only 8.3% to Reactome. The metabolism-curated panel
    maps 91.6% (1,118/1,220) and is what BiGG-graph pre-GIZMO work used.
    File contains both GSM (GSE129705 GEO) and HSA.* (CorEvitas internal V0)
    samples; we filter to HSA.* and use the V0 visit only.

    V3 (3-month post-treatment) is a fundamentally different biological state
    (post-therapy on-treatment biology) vs V0 (pre-therapy baseline used as
    response-prediction substrate). Mixing them would conflate two distinct
    questions and pseudoreplicate at the subject level. V3 is out of scope
    for this response-prediction analysis.

    Labels: Poor responders → "active"; Good responders → "control"; Moderate
    dropped (ambiguous). Returns (trans_data, None, y_label, common).
    """
    import pandas as pd
    METAB_VST = Path("/home/jgardner/gitlab-old/c17edaae86e4016a583e098582f6dbf3eccade8ef83747df9ba617ded9d31309/data/corevitas_gse129705_amdbnorm_combined.exp_data.txt")
    df = pd.read_csv(METAB_VST, sep=r"\s+", engine="python")
    df.index = [str(s).strip('"') for s in df.index]
    df.columns = [str(c).strip('"') for c in df.columns]
    hsa_mask = df.index.str.startswith("HSA.")
    df_hsa = df[hsa_mask]

    metadata_path = Path("/home/jgardner/ra_cohorts/data/corevitas/corevitas_metadata.csv")
    md = pd.read_csv(metadata_path)
    md = md[md["visit"] == 0.0]  # explicit V0-only filter (safety belt)
    md["subject_visit_key"] = md["subject_visit"].astype(str)
    md = md.set_index("subject_visit_key")

    trans_data: dict[str, dict[str, float]] = {}
    y_label: dict[str, str] = {}
    for sid in df_hsa.index:
        if sid not in md.index:
            continue
        resp = md.loc[sid].get("response")
        if pd.isna(resp) or str(resp) not in ("Good", "Poor"):
            continue
        label = "active" if str(resp) == "Poor" else "control"
        y_label[sid] = label
        trans_data[sid] = {gene: float(val) for gene, val in df_hsa.loc[sid].items()
                            if pd.notna(val)}

    common = sorted(trans_data.keys() & y_label.keys())
    return trans_data, None, y_label, common


def load_idh_glioma():
    """Trautwein et al. 2022 (JCI Insight) — GSE190504 + MTBLS3873 paired multi-omic.

    88 diffuse glioma tissue samples with paired RNA-seq (GSE190504) AND
    per-patient NMR metabolomics (MTBLS3873). IDH1-mut (n=53) vs IDH1-wt
    (n=35). The NMR is loaded from a locally-processed concentration matrix
    derived from the MTBLS3873 raw FIDs via `gizmo.io.nmr` pipeline (see
    `.cache/mtbls3873/processed_metabolites.tsv`). The RNA-seq matrix is in
    `benchmarks/glioma_multiomics/glioma_processed.xlsx` (from GSE190504).

    Sample mapping: `description` column in sample_metadata.tsv links NMR ID
    (e.g. "1201_NMR") to rna_id (e.g. "20087P01A01_01").

    Returns (rna_data, metab_data, y_label, common):
        rna_data : {rna_id: {gene_symbol: log2_expression}}
        metab_data : {rna_id: {metabolite_name: tsp-normalized_concentration}}
        y_label : {rna_id: "active" (IDH-mut) | "control" (IDH-wt)}
        common : sorted list of rna_ids with both modalities

    Sources:
      RNA: GEO GSE190504 → benchmarks/glioma_multiomics/glioma_processed.xlsx
      NMR: MetaboLights MTBLS3873 → .cache/mtbls3873/processed_metabolites.tsv
           (processed with gizmo.io.nmr targeted-integration pipeline)
    """
    import pandas as pd
    from pathlib import Path
    glioma_dir = REPO / "benchmarks" / "glioma_multiomics"
    xlsx = glioma_dir / "glioma_processed.xlsx"
    meta = pd.read_csv(glioma_dir / "sample_metadata.tsv", sep="\t")

    raw = pd.read_excel(xlsx, sheet_name=0)
    gene_symbols = raw.iloc[0, 2:].astype(str).tolist()
    sample_rows = raw.iloc[3:, :].copy()
    sample_rows.columns = ["sample_id_a", "rna_id"] + gene_symbols

    rna_to_geno = dict(zip(meta.rna_id.astype(str), meta.genotype))
    # NMR-id → rna_id bridge (so we can index metab by rna_id)
    nmr_to_rna = dict(zip(meta["description"].astype(str),
                              meta["rna_id"].astype(str)))

    # Load per-patient NMR metabolite concentrations
    nmr_path = Path.home() / ".cache" / "mtbls3873" / "processed_metabolites.tsv"
    metab_data: dict[str, dict[str, float]] = {}
    if nmr_path.exists():
        nmr_df = pd.read_csv(nmr_path, sep="\t", index_col=0)
        # Drop tsp_integral (normalization scalar, not a metabolite)
        if "tsp_integral" in nmr_df.columns:
            nmr_df = nmr_df.drop(columns=["tsp_integral"])
        for nmr_id, row in nmr_df.iterrows():
            rna_id = nmr_to_rna.get(str(nmr_id))
            if rna_id is None:
                continue
            metab_data[rna_id] = {m: float(v) for m, v in row.items()
                                       if pd.notna(v) and v > 0}

    rna_data, y_label = {}, {}
    for _, row in sample_rows.iterrows():
        rna_id = str(row["rna_id"])
        if rna_id not in rna_to_geno:
            continue
        geno = rna_to_geno[rna_id]
        if geno not in ("IDH1-mut", "IDH1-wt"):
            continue
        d = {}
        for sym in gene_symbols:
            if not sym or sym == "nan":
                continue
            try:
                v = float(row[sym])
            except (TypeError, ValueError):
                continue
            if pd.isna(v):
                continue
            d[sym] = max(d.get(sym, -1e9), v)
        if not d:
            continue
        rna_data[rna_id] = d
        y_label[rna_id] = "active" if geno == "IDH1-mut" else "control"

    # common = samples with both RNA and metab (and matched genotype)
    common = sorted(set(rna_data.keys()) & set(metab_data.keys()))
    return rna_data, metab_data, y_label, common


def load_tcga_idh_glioma():
    """TCGA-LGG/GBM pan-glioma RNA-seq (n~461) with IDH labels.

    Second IDH glioma cohort for cross-cohort within-disease replication
    (paired with GSE190504 Trautwein 2022 n=88).

    Expression: /home/jgardner/gitlab-old/.../TCGA_GBM_and_LGG_PREPROCESSED_RNASEQ_EXPRESSION.tsv
        672 samples × 15000 genes (gene symbols as columns, log2 expression).
    IDH labels: /home/jgardner/.cache/tcga_idh/lgggbm_tcga_pub_idh_status.tsv
        461 patients from Ceccarelli 2016 (lgggbm_tcga_pub via cBioPortal).
        Built from IDH1/IDH2 somatic mutations.
    Sample-ID format: TCGA.02.0047 (dots) in expression vs TCGA-02-0047
        (dashes) in labels. Normalized to dashes.
    """
    import pandas as pd
    expr_path = Path(
        "/home/jgardner/gitlab-old/d2f483672c0239f6d7dd3c9ecee6deacbcd59185855625902a8b1c1a3bd67440/"
        "TCGA_BRAIN_EXPRESSION/TCGA_GBM_and_LGG_PREPROCESSED_RNASEQ_EXPRESSION.tsv")
    idh_path = Path.home() / ".cache" / "tcga_idh" / "lgggbm_tcga_pub_idh_status.tsv"
    if not (expr_path.exists() and idh_path.exists()):
        raise FileNotFoundError(f"TCGA expr or IDH file missing")

    idh = pd.read_csv(idh_path, sep="\t")
    pid_to_idh = dict(zip(idh.patient_id, idh.idh_status))

    expr = pd.read_csv(expr_path, sep="\t", index_col=0)
    expr.index = [sid.replace(".", "-") for sid in expr.index]

    prot_data, y_label, common = {}, {}, []
    for sid in expr.index:
        if sid not in pid_to_idh:
            continue
        d = {f: float(v) for f, v in expr.loc[sid].items()
              if isinstance(v, (int, float)) and not pd.isna(v)}
        if not d: continue
        prot_data[sid] = d
        y_label[sid] = ("active" if pid_to_idh[sid] == "IDH1-mut"
                          else "control")
        common.append(sid)
    return prot_data, None, y_label, sorted(common)


def load_filbin_covid():
    """Filbin et al. 2021 (Cell Reports Medicine) — MGH COVID Olink cohort.

    Independent COVID cohort for cross-validation against Su_COVID. Same
    platform class (Olink, prot only) but n=383 patients at Day 0 with
    a different lab + protocol. Matches Su's edge_informed design only
    (no metabolomics in Filbin).

    Source: Mendeley Data 10.17632/nf853r8xsj.2
    Files cached under ~/.cache/filbin_mgh_covid/
    """
    import openpyxl
    cache = Path.home() / ".cache" / "filbin_mgh_covid"
    olink_p = cache / "Olink_Proteomics.xlsx"
    clin_p  = cache / "Clinical_Metadata.xlsx"
    assay_p = cache / "Suppl_T2_Olink_Assays_NPX.xlsx"
    if not (olink_p.exists() and clin_p.exists() and assay_p.exists()):
        raise FileNotFoundError(f"Filbin data missing under {cache}")

    # OID → gene symbol (Assay column) from supplemental table 2A
    import pandas as pd
    t2a = pd.read_excel(assay_p, sheet_name="2A-Olink-Assay", header=1)
    # Header is on row 1 because the title takes row 0
    oid_to_sym = dict(zip(t2a["OlinkID"].astype(str),
                            t2a["Assay"].astype(str)))

    # Clinical: Public ID (int) → COVID (0/1)
    clin = pd.read_excel(clin_p, sheet_name="Subject-level metadata")
    pid_to_covid = dict(zip(clin["Public ID"].astype(int).astype(str),
                              clin["COVID"].astype(int)))

    # Olink NPX: filter to Day 0 baseline
    ol = pd.read_excel(olink_p)
    ol = ol[ol.Day == 0].copy()
    ol["pid"] = ol["Public ID"].astype(str).str.replace(r"_D0$", "",
                                                          regex=True)

    prot_data = {}
    y_label = {}
    common = []
    for _, row in ol.iterrows():
        pid = row["pid"]
        if pid not in pid_to_covid:
            continue
        d = {}
        for oid_col, val in row.items():
            if oid_col in ("Public ID", "Day", "pid"):
                continue
            if not isinstance(oid_col, str) or not oid_col.startswith("OID"):
                continue
            sym = oid_to_sym.get(oid_col)
            if not sym or sym == "nan":
                continue
            try:
                v = float(val)
            except (TypeError, ValueError):
                continue
            if pd.isna(v):
                continue
            # Aggregate duplicate gene rows by max NPX (rare)
            d[sym] = max(d.get(sym, -1e9), v)
        if not d:
            continue
        prot_data[pid] = d
        y_label[pid] = "active" if pid_to_covid[pid] == 1 else "control"
        common.append(pid)

    return prot_data, None, y_label, sorted(common)


def load_erawijantari():
    """Load Erawijantari gastrectomy metab matrix (no proteomics)."""
    import csv
    DATA_CANDIDATES = [
        Path("/root/.cache/gemma/demo/curated/ERAWIJANTARI_GASTRIC_CANCER_2020"),
        Path.home() / ".cache/gemma/demo/curated/ERAWIJANTARI_GASTRIC_CANCER_2020",
    ]
    DATA = None
    for d in DATA_CANDIDATES:
        try:
            if d.exists():
                DATA = d
                break
        except PermissionError:
            continue
    if DATA is None:
        raise FileNotFoundError("Erawijantari curated data not found")

    sample_group = {}
    with (DATA / "metadata.tsv").open() as f:
        for row in csv.DictReader(f, delimiter="\t"):
            sample_group[row["Sample"]] = row.get("Study.Group", "").strip()
    metab_data, groups = {}, {}
    with (DATA / "mtb.tsv").open() as f:
        reader = csv.DictReader(f, delimiter="\t")
        all_metabs = [c for c in reader.fieldnames if c != "Sample"]
        for row in reader:
            sid = row["Sample"]
            grp = sample_group.get(sid, "")
            if not grp: continue
            d = {}
            for m in all_metabs:
                v = row[m].strip()
                if not v: continue
                try: val = float(v)
                except ValueError: continue
                if val < 0 or math.isnan(val): continue
                d[m] = math.log2(val + 1e-9)
            metab_data[sid] = d
            groups[sid] = "active" if grp != "Healthy" else "control"

    common = sorted(metab_data.keys() & groups.keys())
    return None, metab_data, groups, common


def load_kmplot_brca():
    """KMPLOT-BRCA breast cancer transcriptomics (Györffy lab meta-cohort).

    Binary label: G3 (high-grade, "active") vs G1/G2 (low-grade, "control").
    Returns (trans_data, None, y_label, common) — single-omic, edge_informed only.
    """
    import pandas as pd
    KMPLOT = Path("/home/jgardner/gitlab-old/d2f483672c0239f6d7dd3c9ecee6deacbcd59185855625902a8b1c1a3bd67440/KMPLOT_BRCA_EXPRESSION")
    xp = pd.read_csv(KMPLOT / "KMPLOT_BRCA_XP_NORMALIZED_CLEANED.tsv", sep="\t")
    xp = xp.rename(columns={xp.columns[0]: "sample_id"})
    xp["sample_id"] = xp["sample_id"].astype(str)
    xp = xp.set_index("sample_id")

    surv = pd.read_csv(KMPLOT / "DATA/KMPLOT_BRCA_SURVIVAL.txt", sep="\t")
    surv["AffyID"] = surv["AffyID"].astype(str)
    surv = surv[surv["Grade"].notna()].copy()
    surv["Grade"] = pd.to_numeric(surv["Grade"], errors="coerce")
    surv = surv[surv["Grade"].notna()].copy()
    grade_dict = dict(zip(surv["AffyID"], surv["Grade"].astype(int)))

    trans_data, y_label = {}, {}
    for sid in xp.index:
        if sid not in grade_dict: continue
        g = grade_dict[sid]
        if g == 3:
            y_label[sid] = "active"  # high-grade
        elif g in (1, 2):
            y_label[sid] = "control"  # low-grade
        else:
            continue
        row = xp.loc[sid]
        trans_data[sid] = {gene: float(v) for gene, v in row.items() if pd.notna(v)}

    common = sorted(set(trans_data.keys()) & set(y_label.keys()))
    return trans_data, None, y_label, common


def load_tcga_luad():
    """TCGA-LUAD lung adenocarcinoma transcriptomics.

    Binary label: stage III/IV ("active") vs stage I/II ("control").
    Returns (trans_data, None, y_label, common) — single-omic.
    """
    import pandas as pd
    LUAD_DIR = Path.home() / ".cache" / "tcga_luad"
    EXPR = LUAD_DIR / "gdac.broadinstitute.org_LUAD.Merge_rnaseqv2__illuminahiseq_rnaseqv2__unc_edu__Level_3__RSEM_genes_normalized__data.Level_3.2016012800.0.0" / "LUAD.rnaseqv2__illuminahiseq_rnaseqv2__unc_edu__Level_3__RSEM_genes_normalized__data.data.txt"
    CLIN = LUAD_DIR / "gdac.broadinstitute.org_LUAD.Clinical_Pick_Tier1.Level_4.2016012800.0.0" / "LUAD.clin.merged.picked.txt"

    # Stage parsing
    def parse_stage(s):
        if not isinstance(s, str): return None
        s = s.lower().strip()
        if "iv" in s: return 4
        if "iii" in s: return 3
        if "ii" in s and "iii" not in s: return 2
        if "i" in s and "ii" not in s and "iii" not in s: return 1
        return None

    cdf = pd.read_csv(CLIN, sep="\t", header=None, low_memory=False)
    attrs = cdf.iloc[:, 0].astype(str).tolist()
    pids = [str(p).strip().lower() for p in cdf.iloc[0, 1:].tolist()]
    if "pathologic_stage" not in attrs:
        return {}, None, {}, []
    stage_idx = attrs.index("pathologic_stage")
    stages = cdf.iloc[stage_idx, 1:].tolist()
    stage_map = {pid: parse_stage(s) for pid, s in zip(pids, stages)
                  if parse_stage(s) is not None}

    edf = pd.read_csv(EXPR, sep="\t", skiprows=[1], low_memory=False)
    edf = edf.rename(columns={"Hybridization REF": "gene"})
    edf["gene"] = edf["gene"].astype(str).str.split("|").str[0]
    edf = edf[edf["gene"].notna() & (edf["gene"] != "?")].drop_duplicates(subset="gene", keep="first")
    edf = edf.set_index("gene")
    new_cols = []
    for c in edf.columns:
        c = c.lower()
        parts = c.split("-")
        if len(parts) >= 3 and parts[0] == "tcga":
            new_cols.append("-".join(parts[:3]))
        else:
            new_cols.append(c)
    edf.columns = new_cols
    edf = edf.loc[:, ~edf.columns.duplicated(keep="first")]
    for c in edf.columns:
        edf[c] = pd.to_numeric(edf[c], errors="coerce")

    trans_data, y_label = {}, {}
    for sid in edf.columns:
        if sid not in stage_map: continue
        stg = stage_map[sid]
        # High-stage (III/IV) = active; Low-stage (I/II) = control
        y_label[sid] = "active" if stg >= 3 else "control"
        col = edf[sid]
        d = {g: float(v) for g, v in col.items() if pd.notna(v) and v > 0}
        if d: trans_data[sid] = d

    common = sorted(set(trans_data.keys()) & set(y_label.keys()))
    return trans_data, None, y_label, common


def load_gse89408_ra():
    """GSE89408 RA-vs-OA synovial tissue RNA-seq (n~174 after filter).

    Binary label: RA tissue ("active") vs OA tissue ("control").
    Returns (trans_data, None, y_label, common).
    """
    import pandas as pd
    GSE = Path.home() / ".cache" / "gse89408" / "GSE89408_count_matrix.txt.gz"
    df = pd.read_csv(GSE, sep="\t", index_col=0)
    df.index.name = "gene"

    def classify(name):
        n = name.lower()
        if n.startswith("ra_tissue"): return "active"
        if n.startswith("oa_tissue"): return "control"
        return None

    keep = [s for s in df.columns if classify(s) is not None]
    sub = df[keep]
    # Log-transform counts (CPM-ish)
    sub = sub.apply(lambda c: np.log2(c.astype(float) + 1))
    trans_data, y_label = {}, {}
    for sid in keep:
        y_label[sid] = classify(sid)
        col = sub[sid]
        trans_data[sid] = {g: float(v) for g, v in col.items() if pd.notna(v) and v > 0}

    common = sorted(set(trans_data.keys()) & set(y_label.keys()))
    return trans_data, None, y_label, common


def load_hmp2_ibd_cd():
    """HMP2 IBDMDB Crohn's Disease metabolomics + biopsy biomarkers.

    Binary label: CD ("active") vs nonIBD ("control"). UC dropped.
    Currently metabolomics-only (the HMP2 directory has metab matrix +
    metadata; biopsy proteomics not in same wrap). Returns
    (None, metab_data, y_label, common) — single-omic node_informed.
    """
    import pandas as pd
    HMP2 = Path("/home/jgardner/GeMMA/data/hmp2")
    metab = pd.read_csv(HMP2 / "hmp2_metabolomics_csm.csv")
    # First col is metabolite ID; remaining cols are sample IDs
    metab.columns = [str(c).strip() for c in metab.columns]
    if metab.columns[0].lower() in ("metabolite", "compound", "id", "name", "unnamed: 0"):
        metab = metab.rename(columns={metab.columns[0]: "metabolite"})
    else:
        metab = metab.rename(columns={metab.columns[0]: "metabolite"})
    metab["metabolite"] = metab["metabolite"].astype(str)
    metab = metab.drop_duplicates(subset="metabolite", keep="first").set_index("metabolite")
    for c in metab.columns:
        metab[c] = pd.to_numeric(metab[c], errors="coerce")

    md = pd.read_csv(HMP2 / "hmp2_metadata.csv", low_memory=False)
    md = md[md["data_type"] == "metabolomics"].copy()
    md = md.drop_duplicates(subset="External ID")
    diag_map = dict(zip(md["External ID"].astype(str), md["diagnosis"].astype(str)))

    metab_data, y_label = {}, {}
    for sid in metab.columns:
        diag = diag_map.get(sid)
        if diag == "CD":
            y_label[sid] = "active"
        elif diag == "nonIBD":
            y_label[sid] = "control"
        else:
            continue
        col = metab[sid]
        d = {m: float(v) for m, v in col.items() if pd.notna(v) and v > 0}
        if d:
            metab_data[sid] = {m: float(np.log2(v + 1e-9)) for m, v in d.items()}

    common = sorted(set(metab_data.keys()) & set(y_label.keys()))
    return None, metab_data, y_label, common


# ---------------------------------------------------------------------------
# Run one (cohort, design) cell
# ---------------------------------------------------------------------------

def run_cell(mg, cohort_name: str, prot_data, metab_data, y_label,
              common: list[str], design: str) -> list[dict]:
    """Run Laplacian + submodular + WL/SP kernels for a given design.

    design ∈ {"edge_informed", "node_informed", "both_directional"}.
    Returns a list of result dicts.
    """
    from per_patient_classification import build_per_patient_subgraphs, graph_kernel_eval
    from gizmo.integration import score_per_sample_laplacian

    # Per-design sample filtering: edge needs prot, node needs metab,
    # both needs intersection
    if design == "edge_informed":
        common = [s for s in common if prot_data and s in prot_data]
    elif design == "node_informed":
        common = [s for s in common if metab_data and s in metab_data]
    elif design == "both_directional":
        common = [s for s in common
                   if prot_data and s in prot_data
                   and metab_data and s in metab_data]
    if not common:
        print(f"  {cohort_name}/{design} — no samples available; skipping")
        return []

    y = np.array([1 if y_label[s] == "active" else 0 for s in common])
    print(f"  {cohort_name}/{design} — {Counter(y_label[s] for s in common)}",
          flush=True)

    # Build / load PubChem name cache for any metab cohort
    pubchem_cache = None
    if metab_data is not None:
        from pathlib import Path as _Path
        from gizmo.evidence.feature_normalize import build_pubchem_cache
        cache_dir = _Path.home() / ".cache" / "gizmo" / "pubchem_name_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = cache_dir / f"{cohort_name}.json"
        all_metab_names = sorted({k for s in common
                                    for k in (metab_data.get(s) or {}).keys()})
        if all_metab_names:
            pubchem_cache = build_pubchem_cache(
                all_metab_names[:300], cache_path,
                rate_limit_s=0.2, max_api_calls=300)

    if design == "edge_informed":
        if prot_data is None:
            print(f"    skipping (no prot data)")
            return []
        per = score_per_sample_laplacian(
            mg, {s: prot_data[s] for s in common},
            feature_assay="proteomics", reference_method="cohort_mean",
            alpha=0.5,
        )
        prot_per, metab_per = per, per
    elif design == "node_informed":
        if metab_data is None:
            print(f"    skipping (no metab data)")
            return []
        per = score_per_sample_laplacian(
            mg, {s: metab_data[s] for s in common},
            feature_assay="metabolomics", reference_method="cohort_mean",
            laplacian_kwargs={"pubchem_name_cache": pubchem_cache},
            alpha=0.5,
        )
        prot_per, metab_per = per, per
    elif design == "both_directional":
        if prot_data is None or metab_data is None:
            print(f"    skipping (need both omics)")
            return []
        prot_per = score_per_sample_laplacian(
            mg, {s: prot_data[s] for s in common},
            feature_assay="proteomics", reference_method="cohort_mean",
            alpha=0.5,
        )
        metab_per = score_per_sample_laplacian(
            mg, {s: metab_data[s] for s in common},
            feature_assay="metabolomics", reference_method="cohort_mean",
            alpha=0.5,
            laplacian_kwargs={"pubchem_name_cache": pubchem_cache},
        )
    else:
        raise ValueError(f"unknown design: {design}")

    SUB_KW = dict(selection="submodular", lam=0.4,
                   min_gain_frac=0.10, min_size=15, max_size=80)
    graphs = build_per_patient_subgraphs(prot_per, metab_per, mg, common, **SUB_KW)
    print(f"    built {len(graphs)} subgraphs", flush=True)

    rows = []
    N_SEEDS = 5
    for kn in ("WL", "SP"):
        try:
            res = graph_kernel_eval(graphs, y, kernel_name=kn, n_seeds=N_SEEDS)
            rows.append({
                "cohort": cohort_name, "design": design, "kernel": kn,
                "n": len(common),
                "n_active": int(y.sum()),
                "n_control": int((1 - y).sum()),
                "cv_auc": res["cv_mean_auc"],
                "cv_sd": res["cv_sd_auc"],
                "cv_seed_sd": res["cv_mean_auc_across_seeds_sd"],
                "loo_auc": res["loo_auc"],
                "loo_seed_sd": res["loo_auc_across_seeds_sd"],
                "holdout_auc": res["holdout_auc"],
                "holdout_seed_sd": res["holdout_auc_across_seeds_sd"],
                "n_seeds": N_SEEDS,
            })
            print(f"    {kn}: CV={res['cv_mean_auc']:.3f}±{res['cv_sd_auc']:.3f} "
                  f"(seed sd {res['cv_mean_auc_across_seeds_sd']:.3f})  "
                  f"LOO={res['loo_auc']:.3f}±{res['loo_auc_across_seeds_sd']:.3f}  "
                  f"hold-out={res['holdout_auc']:.3f}±{res['holdout_auc_across_seeds_sd']:.3f}",
                  flush=True)
        except Exception as exc:
            print(f"    {kn} failed: {exc}", flush=True)
    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 80)
    print("Master per-patient benchmark — submodular + WL/SP × cohorts × designs")
    print("=" * 80)

    print("\nLoading graph...")
    from idh_validation import build_human_graph
    from gizmo.sources.gene_enrichment import enrich_graph_genes_from_reactions
    from gizmo.analysis.currency import (
        flag_currency_metabolites, compute_conditional_currency_edges,
    )
    mg = build_human_graph()
    enrich_graph_genes_from_reactions(mg)
    flag_currency_metabolites(mg, degree_threshold_k=None, include_borderline=False)
    compute_conditional_currency_edges(mg)

    # Allow CLI selection of cohort subset to re-run individual cells
    only_cohort = sys.argv[1] if len(sys.argv) > 1 else None
    cohorts = [
        ("Crohn",    load_crohn,    ["edge_informed", "node_informed", "both_directional"]),
        ("Su_COVID", load_su_covid, ["edge_informed", "node_informed", "both_directional"]),
        ("Erawijantari", load_erawijantari, ["node_informed"]),
        # Gao RA: cohort-paired (different patient sets per omic), so
        # both_directional is excluded — no shared sample IDs.
        ("Gao_RA",   load_gao_ra,   ["edge_informed", "node_informed"]),
        # CorEvitas: trans-only V0 baseline (metabolism-curated 1,220-gene panel),
        # response labels Poor=active vs Good=control (Mod dropped).
        ("CorEvitas", load_corevitas, ["edge_informed"]),
    ]
    if only_cohort:
        cohorts = [c for c in cohorts if c[0] == only_cohort]
        if not cohorts:
            print(f"unknown cohort: {only_cohort}")
            return

    all_rows = []
    for cohort_name, loader, designs in cohorts:
        print(f"\n{'#' * 60}\n# {cohort_name}\n{'#' * 60}")
        try:
            prot_data, metab_data, y_label, common = loader()
            print(f"  loaded: n={len(common)}, classes={Counter(y_label[s] for s in common)}")
        except Exception as exc:
            print(f"  load failed: {exc}")
            continue

        for design in designs:
            try:
                rows = run_cell(mg, cohort_name, prot_data, metab_data,
                                  y_label, common, design)
                all_rows.extend(rows)
            except Exception as exc:
                print(f"  {design} failed: {exc}")

    out_path = RESULTS / "per_patient_master.tsv"
    cols = ["cohort", "design", "kernel", "n", "n_active", "n_control",
            "cv_auc", "cv_sd", "cv_seed_sd", "loo_auc", "loo_seed_sd",
            "holdout_auc", "holdout_seed_sd", "n_seeds"]
    out_path.write_text(
        "\t".join(cols) + "\n" +
        "\n".join("\t".join(str(r[c]) for c in cols) for r in all_rows) + "\n"
    )
    print(f"\n\nWrote: {out_path}")
    print("\nMaster table (mean across 5 seeds; ± is across-seed SD):")
    print(f"{'cohort':<13}{'design':<20}{'kernel':<6}{'n':>5}"
          f"{'CV':>15}{'LOO':>15}{'hold':>15}")
    for r in all_rows:
        print(f"{r['cohort']:<13}{r['design']:<20}{r['kernel']:<6}"
              f"{r['n']:>5}"
              f"  {r['cv_auc']:.3f}±{r['cv_seed_sd']:.3f}"
              f"  {r['loo_auc']:.3f}±{r['loo_seed_sd']:.3f}"
              f"  {r['holdout_auc']:.3f}±{r['holdout_seed_sd']:.3f}")


if __name__ == "__main__":
    main()
