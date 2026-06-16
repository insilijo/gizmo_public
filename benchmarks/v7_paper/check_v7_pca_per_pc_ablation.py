"""v7 Phase 5a: per-PC ablation for BOTH F-α-PCs AND PCA-on-input.

For each cohort, decompose the input into top-7 PCs via straight PCA, then
test each PC individually as a Cox/logistic-regression predictor of
survival/mortality. Compare per-PC discrimination to the F-α-PC ablation
already in v7_survival_ablation.json + v7_survival_tcga_idh.json + the
Filbin script.

Question: does PCA-on-input show the same "PC3 wins" pattern? If yes, then
PC3 isn't substrate-coordinate-specific — it's a generic feature of
unsupervised projection ordering. If no, α-PC3 carries something the
substrate-projection uniquely surfaces.
"""
from __future__ import annotations
import sys, json, time
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
from sklearn.decomposition import PCA
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import roc_auc_score

REPO = Path('/home/jgardner/GIZMO')
RESULTS = REPO / 'benchmarks/results/unsupervised'
ZS_DIR = RESULTS / 'zscored'
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(REPO / 'benchmarks'))


def cv_metric(X, y_or_time_event, label, kind='auc', n_splits=5, seed=42):
    """5-fold CV: returns mean+std of AUC (kind='auc') or C-index (kind='cox')."""
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=seed)
    scores = []
    for tr, te in kf.split(X):
        if X.ndim == 1:
            Xt = X.reshape(-1, 1)
        else:
            Xt = X
        sc = StandardScaler()
        Xs_tr = sc.fit_transform(Xt[tr]); Xs_te = sc.transform(Xt[te])
        if kind == 'auc':
            y = y_or_time_event
            clf = LogisticRegression(max_iter=1000, C=1.0)
            try:
                clf.fit(Xs_tr, y[tr])
                proba = clf.predict_proba(Xs_te)[:, 1]
                scores.append(roc_auc_score(y[te], proba))
            except (ValueError, Exception):
                continue
        elif kind == 'cox':
            from lifelines import CoxPHFitter
            from lifelines.utils import concordance_index
            time_arr, event_arr = y_or_time_event
            df = pd.DataFrame(Xs_tr,
                                columns=[f'f{i}' for i in range(Xt.shape[1])])
            df['time'] = time_arr[tr]; df['event'] = event_arr[tr]
            cph = CoxPHFitter(penalizer=0.01)
            try:
                cph.fit(df, duration_col='time', event_col='event',
                        show_progress=False)
            except Exception:
                continue
            df_te = pd.DataFrame(Xs_te,
                                   columns=[f'f{i}' for i in range(Xt.shape[1])])
            rs = cph.predict_partial_hazard(df_te).values
            scores.append(concordance_index(time_arr[te], -rs, event_arr[te]))
    if not scores: return None
    return float(np.mean(scores)), float(np.std(scores))


def per_pc_ablation(cohort_name, X_input, y_or_time_event, kind='auc'):
    """Compute PCA-top-7 from X_input, then Cox or AUC CV on each PC alone."""
    # Impute NaN with col mean
    X = np.where(np.isnan(X_input), np.nanmean(X_input, axis=0)[None, :], X_input)
    pca = PCA(n_components=7, random_state=42)
    scores_full = pca.fit_transform(X)
    print(f'\n=== {cohort_name} PCA-on-input per-PC ablation ({kind.upper()}) ===',
          flush=True)
    print(f'  EVR top-7: {pca.explained_variance_ratio_.round(3)}', flush=True)
    print(f'  {"PC":<12}  Mean    ±std', flush=True)
    res = {}
    for k in range(7):
        score = scores_full[:, k]
        r = cv_metric(score, y_or_time_event, f'PC{k+1}', kind=kind)
        if r:
            res[f'PC{k+1}'] = {'mean': r[0], 'std': r[1]}
            print(f'  PC{k+1:<10}  {r[0]:.3f}  ±{r[1]:.3f}', flush=True)
        else:
            res[f'PC{k+1}'] = None
            print(f'  PC{k+1:<10}  FAILED', flush=True)
    return res


def load_kmplot():
    KMPLOT_XP = Path('/home/jgardner/gitlab-old/'
                      'd2f483672c0239f6d7dd3c9ecee6deacbcd59185855625902a8b1c1a3bd67440/'
                      'KMPLOT_BRCA_EXPRESSION/KMPLOT_BRCA_XP_NORMALIZED_CLEANED.tsv')
    KMPLOT_SURV = Path('/home/jgardner/gitlab-old/'
                        'd2f483672c0239f6d7dd3c9ecee6deacbcd59185855625902a8b1c1a3bd67440/'
                        'KMPLOT_BRCA_EXPRESSION/DATA/KMPLOT_BRCA_SURVIVAL.txt')
    xp = pd.read_csv(KMPLOT_XP, sep='\t')
    xp = xp.rename(columns={xp.columns[0]: 'sample_id'})
    xp['sample_id'] = xp['sample_id'].astype(str)
    xp = xp.set_index('sample_id')
    surv = pd.read_csv(KMPLOT_SURV, sep='\t')
    surv['AffyID'] = surv['AffyID'].astype(str)
    surv = surv.dropna(subset=['Death_event (1=death)', 'Death_time'])
    surv['event'] = surv['Death_event (1=death)'].astype(int)
    surv['time'] = pd.to_numeric(surv['Death_time'], errors='coerce')
    surv = surv.dropna(subset=['time'])
    surv = surv[surv['time'] > 0]
    common = sorted(set(surv['AffyID']) & set(xp.index))
    surv = surv.set_index('AffyID').loc[common]
    xp = xp.loc[common]
    X = xp.values.astype(float)
    return X, (surv['time'].values, surv['event'].values.astype(int))


def load_tcga_luad():
    CLIN = Path.home() / '.cache' / 'tcga_luad' / 'gdac.broadinstitute.org_LUAD.Clinical_Pick_Tier1.Level_4.2016012800.0.0' / 'LUAD.clin.merged.picked.txt'
    EXPR = Path.home() / '.cache' / 'tcga_luad' / 'gdac.broadinstitute.org_LUAD.Merge_rnaseqv2__illuminahiseq_rnaseqv2__unc_edu__Level_3__RSEM_genes_normalized__data.Level_3.2016012800.0.0' / 'LUAD.rnaseqv2__illuminahiseq_rnaseqv2__unc_edu__Level_3__RSEM_genes_normalized__data.data.txt'
    cdf = pd.read_csv(CLIN, sep='\t', header=None, low_memory=False)
    attrs = cdf.iloc[:, 0].astype(str).tolist()
    pids = [str(p).strip().lower() for p in cdf.iloc[0, 1:].tolist()]
    vital = [str(v).strip() for v in cdf.iloc[attrs.index('vital_status'), 1:].tolist()]
    dtd = [str(v).strip() for v in cdf.iloc[attrs.index('days_to_death'), 1:].tolist()]
    dtf = [str(v).strip() for v in cdf.iloc[attrs.index('days_to_last_followup'), 1:].tolist()]
    rows = []
    for i, pid in enumerate(pids):
        try: v = int(float(vital[i]))
        except (ValueError, IndexError): continue
        if v == 1:
            try: t = float(dtd[i]); ev = 1
            except (ValueError, TypeError): continue
        elif v == 0:
            try: t = float(dtf[i]); ev = 0
            except (ValueError, TypeError): continue
        else: continue
        if t > 0 and np.isfinite(t):
            rows.append({'patient_id': pid, 'time': t, 'event': ev})
    surv = pd.DataFrame(rows)
    expr = pd.read_csv(EXPR, sep='\t', skiprows=[1])
    expr = expr.set_index(expr.columns[0])
    expr.columns = [c.lower()[:12] if len(c) >= 12 else c.lower()
                    for c in expr.columns]
    expr_t = expr.T.groupby(level=0).mean()
    common = sorted(set(surv['patient_id']) & set(expr_t.index))
    surv = surv.set_index('patient_id').loc[common]
    expr_t = expr_t.loc[common]
    X = np.log2(np.maximum(expr_t.values.astype(float), 0.0) + 1.0)
    return X, (surv['time'].values, surv['event'].values.astype(int))


def load_tcga_idh():
    CLIN = Path.home() / '.cache' / 'tcga_idh' / 'lgggbm_tcga_pub_clinical.tsv'
    EXPR = Path('/home/jgardner/gitlab-old/'
                 'd2f483672c0239f6d7dd3c9ecee6deacbcd59185855625902a8b1c1a3bd67440/'
                 'TCGA_BRAIN_EXPRESSION/TCGA_GBM_and_LGG_PREPROCESSED_RNASEQ_EXPRESSION.tsv')
    clin = pd.read_csv(CLIN, sep='\t')
    clin = clin.dropna(subset=['OS_MONTHS', 'OS_STATUS'])
    clin['event'] = clin['OS_STATUS'].astype(str).str.startswith('1').astype(int)
    clin['time'] = pd.to_numeric(clin['OS_MONTHS'], errors='coerce')
    clin = clin.dropna(subset=['time'])
    clin = clin[clin['time'] > 0]
    clin['patient_id'] = clin['patient_id'].astype(str).str.lower()
    expr = pd.read_csv(EXPR, sep='\t', index_col=0)
    if expr.shape[0] > expr.shape[1]:
        expr.columns = [c.lower().replace('.', '-') for c in expr.columns]
        expr_t = expr.T
    else:
        expr.index = [i.lower().replace('.', '-') for i in expr.index]
        expr_t = expr
    expr_t.index = [s[:12] if len(s) >= 12 else s for s in expr_t.index]
    expr_t = expr_t.groupby(expr_t.index).mean()
    common = sorted(set(clin['patient_id']) & set(expr_t.index))
    clin = clin.set_index('patient_id').loc[common]
    expr_t = expr_t.loc[common]
    return expr_t.values.astype(float), (clin['time'].values,
                                            clin['event'].values.astype(int))


def load_filbin():
    wb = openpyxl.load_workbook('/home/jgardner/.cache/filbin_mgh_covid/Olink_Proteomics.xlsx',
                                  read_only=True)
    ws = wb['Olink Proteomics']
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    d0 = df[df['Day'] == 0].copy()
    d0['patient_id'] = d0['Public ID'].astype(str).str.replace('_D0', '')
    prot_cols = [c for c in d0.columns if str(c).startswith('OID')]
    wb2 = openpyxl.load_workbook('/home/jgardner/.cache/filbin_mgh_covid/Clinical_Metadata.xlsx',
                                   read_only=True)
    ws2 = wb2['Subject-level metadata']
    crows = list(ws2.iter_rows(values_only=True))
    clin = pd.DataFrame(crows[1:], columns=crows[0])
    clin['Public ID'] = clin['Public ID'].astype(str)
    clin['acuity_28'] = pd.to_numeric(clin['Acuity 28'], errors='coerce')
    clin = clin.dropna(subset=['acuity_28']).copy()
    clin['died_28d'] = (clin['acuity_28'] == 1).astype(int)
    common = sorted(set(d0['patient_id']) & set(clin['Public ID']))
    d0 = d0.set_index('patient_id').loc[common]
    clin = clin.set_index('Public ID').loc[common]
    X = d0[prot_cols].apply(pd.to_numeric, errors='coerce').values
    return X, clin['died_28d'].values.astype(int)


def main():
    print('=== v7 PCA per-PC ablation across 4 cohorts ===', flush=True)
    t0 = time.time()

    print('\n--- KMPLOT_BRCA (Cox C-index) ---', flush=True)
    X, te = load_kmplot()
    print(f'  X: {X.shape}', flush=True)
    res_kmplot = per_pc_ablation('KMPLOT_BRCA', X, te, kind='cox')

    print('\n--- TCGA_LUAD (Cox C-index) ---', flush=True)
    X, te = load_tcga_luad()
    print(f'  X: {X.shape}', flush=True)
    res_luad = per_pc_ablation('TCGA_LUAD', X, te, kind='cox')

    print('\n--- TCGA_IDH_glioma (Cox C-index) ---', flush=True)
    X, te = load_tcga_idh()
    print(f'  X: {X.shape}', flush=True)
    res_idh = per_pc_ablation('TCGA_IDH_glioma', X, te, kind='cox')

    print('\n--- Filbin_COVID (Logistic AUC) ---', flush=True)
    X, y = load_filbin()
    print(f'  X: {X.shape}', flush=True)
    res_filbin = per_pc_ablation('Filbin_COVID', X, y, kind='auc')

    out = ZS_DIR / 'v7_pca_per_pc_ablation.json'
    out.write_text(json.dumps({
        'KMPLOT_BRCA': res_kmplot,
        'TCGA_LUAD': res_luad,
        'TCGA_IDH_glioma': res_idh,
        'Filbin_COVID': res_filbin,
        'compute_seconds': time.time() - t0,
    }, indent=2))
    print(f'\nWrote {out}', flush=True)


if __name__ == '__main__':
    main()
